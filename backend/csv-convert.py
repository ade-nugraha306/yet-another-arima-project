import openpyxl
import pandas as pd
import os

def convert_all_sheets_to_csv(xlsx_file, output_folder="./csv_output"):
    # Buat folder output jika belum ada
    os.makedirs(output_folder, exist_ok=True)

    # Buka file Excel
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)

    # Loop semua sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Ambil semua data dari sheet
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))

        # Konversi ke DataFrame
        df = pd.DataFrame(rows)

        # Nama file CSV berdasarkan nama sheet
        safe_name = sheet_name.replace("/", "_").replace("\\", "_")
        csv_file = os.path.join(output_folder, f"{safe_name}.csv")

        # Simpan sebagai CSV
        df.to_csv(csv_file, index=False, header=False)

        print(f"✔ Sheet '{sheet_name}' berhasil dikonversi -> {csv_file}")

# Contoh penggunaan:
convert_all_sheets_to_csv("AVG 12W & 5W (W1-W40).xlsx")
